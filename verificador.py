from openpyxl import load_workbook
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import keyboard
import pyautogui

#Abre planilha
file = "planilha base.xlsx"
workbook = load_workbook(file)
worksheet= workbook['Dados']
primeiro = 0

#Inicializa o WebDriver do Chrome
driver = webdriver.Chrome()

#Itera as linhas da tabela do excel
for row in worksheet.iter_rows():
   if row[0].value != None and row[0].value != "CPF" and row[2].value == None:
      print(row[0].value)
      driver.get("https://servicos.receita.fazenda.gov.br/Servicos/CPF/ConsultaSituacao/ConsultaPublica.asp")

      #Insere CPF no campo
      campo_cpf = driver.find_element(By.NAME, "txtCPF")
      campo_cpf.send_keys(row[0].value)
        
      if row[1].value != None:
         #Insere data de nascimento no campo
         campo_data = driver.find_element(By.NAME, "txtDataNascimento")
         campo_data.send_keys(row[1].value)
         
         time.sleep(1)
         
         #clica no captcha
         x,y = pyautogui.locateCenterOnScreen('botao_captcha.png')
         pyautogui.click(x, y)
         
         time.sleep(2)
         
         #clica no botão de consultar
         x,y = pyautogui.locateCenterOnScreen('botao_consultar.png')
         pyautogui.click(x, y)
         
         #Pega nome da pessoa
         nome_pessoa = driver.find_element(By.XPATH, '//*[@id="mainComp"]/div[2]/p/span[2]/b').text
         
         #Pega situação da pessoa e insere na planilha
         situacao_cpf = driver.find_element(By.XPATH, '//*[@id="mainComp"]/div[2]/p/span[4]/b').text
         print(situacao_cpf)
         
         #Se o CPF for regular, ele irá salvar o PDF
         if situacao_cpf == "REGULAR":
            #Acessa tela de impressão
            btn_print = driver.find_element(By.ID, "imgPrint")
            btn_print.click()
            
            time.sleep(1)

            x,y = pyautogui.locateCenterOnScreen('botao_imprimir.png')
            pyautogui.click(x, y)
            
            time.sleep(1)
            
            #Se for o primeiro CPF da Lista ele vai precisar mudar o download padrão para o Save to PDF
            if primeiro == 0:
               pyautogui.click(x=1063, y=168)
               keyboard.press_and_release('down')
               time.sleep(1)
               primeiro = 1
         
            keyboard.press_and_release('enter')
            time.sleep(1)

            pyautogui.click(x=998, y=683)

            time.sleep(1)
            #Insere nome do arquivo
            pyautogui.typewrite(nome_pessoa + ".pdf")

            time.sleep(1)
            
            #Salva arquivo
            keyboard.press_and_release('enter')
            
            time.sleep(1)
            
            #fecha janela
            pyautogui.click(x=1263, y=24)
         
         row[2].value = situacao_cpf
         
workbook.save(file)
workbook.close()